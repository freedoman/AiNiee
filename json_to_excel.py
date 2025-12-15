#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
将 JSON 术语表转换为 Excel 文件
用法: python json_to_excel.py <输入json文件路径> [输出excel文件路径]
"""

import json
import sys
import os

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请运行: pip install openpyxl")
    sys.exit(1)


def json_to_excel(json_path, excel_path=None):
    """
    将JSON术语表转换为Excel文件
    
    参数:
        json_path: JSON文件路径
        excel_path: Excel文件输出路径(可选,默认为JSON文件名.xlsx)
    """
    # 如果没有指定输出路径,使用JSON文件名
    if excel_path is None:
        excel_path = os.path.splitext(json_path)[0] + '.xlsx'
    
    # 读取JSON文件
    print(f"正在读取 JSON 文件: {json_path}")
    try:
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
    except FileNotFoundError:
        print(f"错误: 找不到文件 {json_path}")
        return False
    except json.JSONDecodeError as e:
        print(f"错误: JSON 格式错误 - {e}")
        return False
    
    if not isinstance(data, list):
        print("错误: JSON 必须是数组格式")
        return False
    
    print(f"共读取 {len(data)} 条术语")
    
    # 创建Excel工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "术语表"
    
    # 定义表头
    headers = ["原文 (src)", "译文 (dst)", "描述 (info)", "频次 (count)"]
    
    # 设置表头样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 写入表头
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # 写入数据
    print("正在写入数据到 Excel...")
    for row_idx, item in enumerate(data, start=2):
        # 原文
        cell = ws.cell(row=row_idx, column=1)
        cell.value = item.get('src', '')
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        
        # 译文
        cell = ws.cell(row=row_idx, column=2)
        cell.value = item.get('dst', '')
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        
        # 描述
        cell = ws.cell(row=row_idx, column=3)
        cell.value = item.get('info', '')
        cell.border = border
        cell.alignment = Alignment(vertical="center")
        
        # 频次
        cell = ws.cell(row=row_idx, column=4)
        count = item.get('count', 0)
        cell.value = count if isinstance(count, int) else 0
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # 自动调整列宽
    for col in range(1, 5):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col)
        
        for cell in ws[column_letter]:
            try:
                if cell.value:
                    # 计算字符串长度(中文按2个字符计算)
                    cell_len = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                    if cell_len > max_length:
                        max_length = cell_len
            except:
                pass
        
        # 设置列宽(最小10,最大50)
        adjusted_width = min(max(max_length + 2, 10), 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 添加自动筛选
    ws.auto_filter.ref = ws.dimensions
    
    # 保存Excel文件
    print(f"正在保存 Excel 文件: {excel_path}")
    try:
        wb.save(excel_path)
        print(f"✓ 转换成功!")
        print(f"  输出文件: {excel_path}")
        print(f"  共 {len(data)} 条术语")
        return True
    except Exception as e:
        print(f"错误: 保存文件失败 - {e}")
        return False


def main():
    """主函数"""
    # 检查命令行参数
    if len(sys.argv) < 2:
        print("用法: python json_to_excel.py <输入json文件路径> [输出excel文件路径]")
        print("\n示例:")
        print("  python json_to_excel.py 导出_术语表.json")
        print("  python json_to_excel.py 导出_术语表.json 术语表.xlsx")
        sys.exit(1)
    
    json_path = sys.argv[1]
    excel_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    # 执行转换
    success = json_to_excel(json_path, excel_path)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
