#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
将 Excel 术语表转换回 JSON 文件
用法: python excel_to_json.py <输入excel文件路径> [输出json文件路径]
"""

import json
import sys
import os

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("请运行: pip install openpyxl")
    sys.exit(1)


def excel_to_json(excel_path, json_path=None):
    """
    将Excel术语表转换为JSON文件
    
    参数:
        excel_path: Excel文件路径
        json_path: JSON文件输出路径(可选,默认为Excel文件名.json)
    """
    # 如果没有指定输出路径,使用Excel文件名
    if json_path is None:
        json_path = os.path.splitext(excel_path)[0] + '.json'
    
    # 读取Excel文件
    print(f"正在读取 Excel 文件: {excel_path}")
    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
    except FileNotFoundError:
        print(f"错误: 找不到文件 {excel_path}")
        return False
    except Exception as e:
        print(f"错误: 无法读取 Excel 文件 - {e}")
        return False
    
    # 检查表头
    headers = [cell.value for cell in ws[1]]
    print(f"检测到表头: {headers}")
    
    # 确定列索引
    col_mapping = {}
    for idx, header in enumerate(headers, start=1):
        if header:
            header_lower = str(header).lower()
            if 'src' in header_lower or '原文' in header_lower:
                col_mapping['src'] = idx
            elif 'dst' in header_lower or '译文' in header_lower:
                col_mapping['dst'] = idx
            elif 'info' in header_lower or '描述' in header_lower:
                col_mapping['info'] = idx
            elif 'count' in header_lower or '频次' in header_lower or '次数' in header_lower:
                col_mapping['count'] = idx
    
    # 检查必需的列是否存在
    required_cols = ['src', 'dst', 'info', 'count']
    missing_cols = [col for col in required_cols if col not in col_mapping]
    if missing_cols:
        print(f"警告: 缺少以下列: {missing_cols}")
        print("将使用默认值填充缺失的字段")
    
    # 读取数据
    data = []
    row_count = 0
    
    print("正在读取数据...")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # 跳过空行
        if not any(row):
            continue
        
        # 构建术语条目
        item = {}
        
        # 原文
        if 'src' in col_mapping:
            src = row[col_mapping['src'] - 1]
            item['src'] = str(src).strip() if src is not None else ""
        else:
            item['src'] = ""
        
        # 译文
        if 'dst' in col_mapping:
            dst = row[col_mapping['dst'] - 1]
            item['dst'] = str(dst).strip() if dst is not None else ""
        else:
            item['dst'] = ""
        
        # 描述
        if 'info' in col_mapping:
            info = row[col_mapping['info'] - 1]
            item['info'] = str(info).strip() if info is not None else ""
        else:
            item['info'] = ""
        
        # 频次
        if 'count' in col_mapping:
            count = row[col_mapping['count'] - 1]
            try:
                item['count'] = int(count) if count is not None else 1
            except (ValueError, TypeError):
                item['count'] = 1
        else:
            item['count'] = 1
        
        # 只添加有原文的条目
        if item['src']:
            data.append(item)
            row_count += 1
    
    print(f"共读取 {row_count} 条术语")
    
    if row_count == 0:
        print("警告: 没有读取到任何数据")
        return False
    
    # 保存JSON文件
    print(f"正在保存 JSON 文件: {json_path}")
    try:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        
        print(f"✓ 转换成功!")
        print(f"  输出文件: {json_path}")
        print(f"  共 {row_count} 条术语")
        
        # 显示前几条数据作为验证
        if data:
            print("\n前3条数据预览:")
            for i, item in enumerate(data[:3], 1):
                print(f"  {i}. {item}")
        
        return True
    except Exception as e:
        print(f"错误: 保存文件失败 - {e}")
        return False


def main():
    """主函数"""
    # 检查命令行参数
    if len(sys.argv) < 2:
        print("用法: python excel_to_json.py <输入excel文件路径> [输出json文件路径]")
        print("\n示例:")
        print("  python excel_to_json.py 术语表.xlsx")
        print("  python excel_to_json.py 术语表.xlsx 导入_术语表.json")
        sys.exit(1)
    
    excel_path = sys.argv[1]
    json_path = sys.argv[2] if len(sys.argv) > 2 else None
    
    # 执行转换
    success = excel_to_json(excel_path, json_path)
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
