import openpyxl
import rapidjson as json
from qfluentwidgets import TableWidget
from PyQt5.QtWidgets import QTableWidgetItem

class TableHelper():

    def __init__(self) -> None:
        super().__init__()

    # 从表格加载数据
    def load_from_table(table: TableWidget, keys: list[str]) -> list[dict]:
        from PyQt5.QtCore import Qt
        result = []

        # 遍历每一行
        for row in range(table.rowCount()):
            # 获取当前行所有条目
            data: list[QTableWidgetItem] = [
                table.item(row, col)
                for col in range(table.columnCount())
            ]

            # 检查数据合法性
            if not isinstance(data[0], QTableWidgetItem) or len(data[0].text().strip()) == 0:
                continue

            # 添加数据
            row_data = {}
            for i in range(len(keys)):
                if isinstance(data[i], QTableWidgetItem):
                    # 对于 count 字段，尝试转换为整数
                    if keys[i] == "count":
                        try:
                            # 尝试从 EditRole 获取数字值
                            edit_value = data[i].data(Qt.EditRole)
                            if edit_value is not None and edit_value != "":
                                row_data[keys[i]] = int(edit_value) if isinstance(edit_value, (int, float)) else int(str(edit_value))
                            else:
                                # 如果没有 EditRole，尝试从文本解析
                                text = data[i].text().strip()
                                row_data[keys[i]] = int(text) if text.isdigit() else 1
                        except (ValueError, TypeError):
                            row_data[keys[i]] = 1  # 默认值
                    else:
                        row_data[keys[i]] = data[i].text().strip()
                else:
                    row_data[keys[i]] = ""
            
            result.append(row_data)

        return result

    # 向表格更新数据
    def update_to_table(table: TableWidget, data: list[dict], keys: list[str]) -> None:
        from PyQt5.QtCore import Qt
        
        # 设置表格行数
        table.setRowCount(max(16, len(data)))

        # 去重
        data_unique = {v.get(keys[0], ""): v for v in data}
        data = [v for v in data_unique.values()]

        # 遍历表格
        for row, v in enumerate(data):
            for col in range(table.columnCount()):
                value = v.get(keys[col], "")
                item = QTableWidgetItem()
                
                # 如果是数字类型字段（如 count），设置为数字以便正确排序
                if keys[col] == "count" and isinstance(value, (int, float)):
                    item.setData(Qt.DisplayRole, str(value))
                    item.setData(Qt.EditRole, value)
                else:
                    item.setText(str(value))
                
                table.setItem(row, col, item)

    # 从文件加载数据
    def load_from_file(path: str, keys: list[str]) -> list[dict]:
        result = []

        # 从 json 文件加载数据
        if path.endswith(".json"):
            result = TableHelper.load_from_json_file(path, keys)

        # 从 xlsx 文件加载数据
        if path.endswith(".xlsx"):
            result = TableHelper.load_from_xlsx_file(path, keys)

        return result

    # 从 json 文件加载数据
    def load_from_json_file(path: str, keys: list[str]) -> list[dict]:
            result = []

            # 读取文件
            inputs = []
            with open(path, "r", encoding = "utf-8") as reader:
                inputs = json.load(reader)

            # 标准字典列表
            # [
            #     {
            #         "key": "value",
            #         "key": "value",
            #         "key": "value",
            #     }
            # ]
            if isinstance(inputs, list):
                for data in inputs:
                    # 数据校验
                    if not isinstance(data, dict) or str(data.get(keys[0], "")).strip() == "":
                        continue

                    # 添加数据
                    result.append({
                        keys[i]: str(data.get(keys[i], "")).strip()
                        for i in range(len(keys))
                    })

                # 兼容旧版，保留一段时间用以过度
                if len(result) == 0 and "src" in keys:
                    # 将 keys 中的 "src" 替换为 "srt" 然后重试
                    result = TableHelper.load_from_json_file(path, [("srt" if v == "src" else v) for v in keys])

                    # 将字段换回来
                    for v in result:
                        v["src"] = v.get("srt", "")

            # 标准 KV 字典
            # [
            #     "ダリヤ": "达莉雅"
            # ]
            if isinstance(inputs, dict):
                for k, v in inputs.items():
                    # 数据校验
                    if str(k).strip() == "":
                        continue

                    # 添加数据
                    item = {}
                    for i in range(len(keys)):
                        if i == 0:
                            item[keys[i]] = str(k).strip()
                        elif i == 1:
                            item[keys[i]] = str(v).strip() if v != None else ""
                        else:
                            item[keys[i]] = ""
                    result.append(item)

            return result

    # 从 xlsx 文件加载数据
    def load_from_xlsx_file(path: str, keys: list[str]) -> list[dict]:
        result = []

        sheet = openpyxl.load_workbook(path).active
        for row in range(2, sheet.max_row + 1): # 跳过标题行，从第二行开始
            # 读取每一行的数据
            data: list[str] = [
                sheet.cell(row = row, column = col).value
                for col in range(1, len(keys) + 1)
            ]

            # 检查数据合法性
            if data[0] == None or str(data[0]).strip() == "":
                continue

            # 添加数据
            result.append(
                {
                    keys[i]: str(data[i]).strip() if data[i] != None else ""
                    for i in range(len(keys))
                }
            )

        return result